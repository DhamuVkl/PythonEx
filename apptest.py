# print("Hi Dhamu")
# print("*" * 20)
# msg0 = 77
# print(msg0)
# name = input('What is your Name? ')
# print('Hi ' + name)
# color = input('What is your favourite color?')
# print(name + ' likes ' + color)
# weight = input('Enter your weight in lb: ')
# weight_in_kg = ((1/(2.20462262)) * int(weight))
# print('Your weight in Kg: ', float(weight_in_kg))
# first = 'Dharageswaran'
# last = 'S'
# msg = first + ' ['+ last +']' + ' is a coder'
# msg1 = f'{first} [{last}] is a coder'  # f'-> formatted string
# print(msg1)  
# print(msg)
# print(len(msg))
# print(msg.upper())
# print(msg)
# print(msg.find('g'))  
# print(msg.replace('Dharageswaran', 'Dharagesh'))
# print('Dharagesh' in msg)   # Boolean check
# print('Dharageswaran' in msg)
# print(10/3)
# print(10//3)
# print(10%3)
# print(2**3)
# x = 9
# x += 3
# print(x)
# x = 5.57
# print(round(x))
# y = -6.76
# print(abs(y))

# import math
# print(math.ceil(3.9))
# print(math.floor(7.9))

# hot = False
# cold = False
# if hot:
#     print('Drink plenty of water')
# elif cold:
#     print('Wear warm clothes')
# else:
#     print('Enjoy your day')


# has_good_credit = True
# has_high_income = True
# has_criminal_record = True
# if has_good_credit and has_high_income and not has_criminal_record:
#     print("Eligible for loan ")
# else:
#     print("Not Eligible for loan")



# msg2 = input('Enter your name: ')
# #msg2 ='dharageswaran'
# print(len(msg2))
# if len(msg2) < 3:
#     print('Name must at least 3 characters')
# elif len(msg2) > 50:
#     print("Name can't  max 50 characters")
# else:
#     print('Name looks good') 
    
# weight1 = int(input("Enter your weight: ")) # input() always store in str type
# unit = input("(L)bs or (K)g: ")

# if unit.lower() == "k":
#     weight_lbs = weight1 * 2.20462262
#     print(f"you are {weight_lbs} pounds")
# if unit.lower() == "l":
#     weight_kg = weight1 / 2.20462262
#     print(f"you are {weight_kg} Kg")

# i = 1
# while i <= 3:
#     value = int(input("Guess 1-9: "))  
#     if value == 7:
#         print("you win!")
#         break
#     else:
#         i= i+1
        
# else:
#      print("you loss") 



# started = False
# while True:     # loops always until breaks
#     info = input(">> ").lower()
#     if info == "help":
#         print("start - to start the car \nstop - to stop the car \nquit - to exit")
#     elif info == "start":
#         if started:
#             print("Hey! car is already started")
#         else:
#             started = True
#             print("Car started...Ready to go!")
#     elif info == "stop":
#         if not started:
#             print("Hey! car is already stopped")
#         else:
#             started = False
#             print("Car is stopped")
#     elif info == "quit":
#         break                               
#     else:
#         print("I don't understand that...")    
    

# prices = [10, 20, 30]

# total = 0
# for i in prices:
#     total += i
# print(f"Total price: {total}")


# numbers = [5, 2, 5, 2, 2]
#                             # nested loop program
# for x_count in numbers:
#     output = ""             # to clear 
#     for count in range(x_count):
#         output += 'x'
#     print(output) 
    

# lists = [1, 23, 10, 17, 9]
# high = lists[0]    
# for large in lists:
#     if large > high:
#         high = large
# print(high)

# lists.insert(0,5)
# lists.append(10)
# print(lists)


# from email import message


# duplicate = [3, 5, 2, 3, 6, 7]
# unique = []
# for number in duplicate:
#     if number not in unique:
#         unique.append(number)
# print(unique)

# coordinates = (1, 2, 3)      # (1, 2, 3)=> are called tuples you can't change values
# x, y, z = coordinates        # called unpacking
# print(y)


# matrix = [                  # 3x3 matrix => list in a list
#     [1, 2, 3],
#     [4, 5, 6],
#     [7, 8, 9]
# ]
# print(matrix[0][1])


# customer = {                 # Dictionaries
#     "name": "Dhamu",         # each key is unique no duplicate allowed
#     "age": 22,
#     "is_verified": True
# }
# customer["birth_date"] = "march 6 2000"
# print(customer["name"])      # print(customer.get(name, ""))    both same output but here no error
# print(customer["birth_date"])



# phone = input("phone: ")
# digits = {
#     "1": "one",
#     "2": "two",
#     "3": "three",
#     "4": "four",
#     "5": "five",
#     "6": "six",
#     "7": "seven",
#     "8": "eight",
#     "9": "nine",
#     "0": "zero"
# }
# output = ""
# for ch in phone:
#     output += digits.get(ch) + " "
# print(output)


# message = input(">> ")
# words = message.split(" ")      # Ex:  hi world :) => ["hi", "world", ":)"] it split by spaces and store in list
# emojis = {
#     ":)": "😊",
#     ":(": "🙁"
# }
# output = ""
# for word in words:
#     output += emojis.get(word, word) + " "         # if word not in dictionaries it put same word
# print(output)




# def user():                 # functions
#     print("Hi !")
#     print("Welcome aboard")


# print("Start")
# user()
# print("Finish")



# def user(first_name, last_name):
#     print(f"Hi {first_name} {last_name}!")
#     print("Welcome aboard")


# print("Start")
# user("Dhamu", "S")
# print("Finish")




# def square(number):            
#     return number*number             # returns value to function
  
  
# print(square(3))




# def emoji_converter(message):       
#     words = message.split(" ")      # Ex:  hi world :) => ["hi", "world", ":)"] it split by spaces and store in list
#     emojis = {
#         ":)": "😊",
#         ":(": "🙁"
#     }
#     output = ""
#     for word in words:
#         output += emojis.get(word, word) + " "
#     return output


# message = input(">> ")
# emoji_converter(message)            # function call
# print(emoji_converter(message))


# try:                                # this used to handle error without crashing the program
#     age = int(input("Age: "))
#     print(age)
# except ValueError:
#     print("Invalid value")
    

# class Point:                    # to define "type" like num, str, bool
#     def move(self):
#         print("move")

#     def draw(self):
#         print("draw")

# point = Point()
# point.draw()


# class Apple:
#     def __init__(self, x, y):
#         self.x = x
#         self.y = y
#         pass


# count = Apple(10, 30)
# print(count.x)


# class Person:
#     def __init__(self, name):
#         self.name = name
#         pass

#     def talk(self):
#         print(f"Hi I'm {self.name}")


# user = Person("Dhamu")
# print(user.name)
# user.talk()
# user_1 = Person("Monika")
# user_1.talk()





# import utils                    # importing module
# lists = [3, 5, 7, 1, 9, 4]
# print(utils.find_max(lists))    # function call



# from ecommerce.shipping import cal_shipping         # package managing
# cal_shipping()





# import random

# class Dice:                     # rolling dice in random
#     def roll(self):
#         x = random.randint(1, 6)
#         y = random.randint(1, 6)
#         return (x, y)


# dice = Dice()
# print(dice.roll())







# from pathlib import Path            # finding path files

# path = Path()
# for file in path.glob('*.py'):
#     print(file)







import openpyxl as xl                           # calling it as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")      # wb = work book
sheet = wb["Sheet1"]


for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,                       # for BarChart
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "e2")
wb.save("transactions_1.xlsx")

    
    




